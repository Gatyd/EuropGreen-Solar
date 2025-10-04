"""
Vues pour la génération de rapports.
Exports comptables, rapports de ventes, commissions, prospects.
"""

import csv
import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Count, Sum, Avg, F, Value, DecimalField
from django.db.models.functions import Coalesce
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiParameter

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Import des modèles
from invoices.models import Invoice
from billing.models import Quote
from request.models import ProspectRequest
from installations.models import Form
from users.models import User


class ReportsViewSet(viewsets.GenericViewSet):
    """
    ViewSet pour la génération de rapports et exports.
    """
    # DRF nécessite un queryset pour le router, même si on n'utilise que des @action
    queryset = Invoice.objects.none()
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
    
    def _validate_date_string(self, date_str):
        """
        Valide et nettoie une chaîne de date.
        Empêche les dates invalides comme '12121-01-01'.
        """
        if not date_str:
            return None
        
        try:
            # Parser la date
            parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
            
            # Vérifier que l'année est raisonnable (entre 1900 et 2100)
            if parsed_date.year < 1900 or parsed_date.year > 2100:
                return None
            
            # Vérifier que la date n'est pas dans le futur
            if parsed_date > datetime.now():
                return None
            
            return parsed_date
        except (ValueError, OverflowError):
            return None
    
    def _get_date_range(self, request):
        """Extrait la période depuis les paramètres de requête."""
        start_str = request.query_params.get('start_date')
        end_str = request.query_params.get('end_date')
        
        if start_str and end_str:
            start_date_parsed = self._validate_date_string(start_str)
            end_date_parsed = self._validate_date_string(end_str)
            
            # Si les dates sont valides, les utiliser
            if start_date_parsed and end_date_parsed:
                start_date = timezone.make_aware(start_date_parsed)
                end_date = timezone.make_aware(end_date_parsed.replace(hour=23, minute=59, second=59))
            else:
                # Dates invalides, utiliser la période par défaut
                end_date = timezone.now()
                start_date = end_date.replace(year=end_date.year - 1)
        else:
            # Par défaut : 1 an
            end_date = timezone.now()
            start_date = end_date.replace(year=end_date.year - 1)
        
        return start_date, end_date
    
    @extend_schema(
        summary="Export comptable",
        description="Export des factures au format CSV ou Excel pour intégration comptable.",
        parameters=[
            OpenApiParameter(name='start_date', description='Date début (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='end_date', description='Date fin (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='status', description='Statut factures (all, paid, issued, partially_paid)', required=False, type=str),
            OpenApiParameter(name='export_format', description='Format export (csv, excel)', required=False, type=str),
        ]
    )
    @action(detail=False, methods=['get'], url_path='accounting-export')
    def accounting_export(self, request):
        """
        Export comptable des factures.
        
        Colonnes exportées :
        - Numéro facture
        - Date émission
        - Date échéance
        - Client (nom complet)
        - Montant HT
        - Taux TVA (%)
        - Montant TVA (€)
        - Montant TTC
        - Statut
        - Notes
        """
        start_date, end_date = self._get_date_range(request)
        status_filter = request.query_params.get('status', 'all')
        export_format = request.query_params.get('export_format', 'csv')
        
        # Récupérer les factures
        invoices = Invoice.objects.filter(
            created_at__range=[start_date, end_date]
        ).select_related('installation__client')
        
        # Filtrer par statut si demandé
        if status_filter != 'all':
            if status_filter == 'paid':
                invoices = invoices.filter(status='paid')
            elif status_filter == 'issued':
                invoices = invoices.filter(status='issued')
            elif status_filter == 'partially_paid':
                invoices = invoices.filter(status='partially_paid')
        
        invoices = invoices.order_by('issue_date')
        
        # Préparer les données
        data = []
        for inv in invoices:
            client = inv.installation.client if inv.installation and inv.installation.client else None
            client_name = f"{client.last_name} {client.first_name}" if client else "Client inconnu"
            
            # Calculer TVA
            subtotal = float(inv.subtotal)
            tax_rate = float(inv.tax_rate)
            tax_amount = subtotal * (tax_rate / 100)
            total = float(inv.total)
            
            # Statut en français
            status_labels = {
                'draft': 'Brouillon',
                'issued': 'Émise',
                'partially_paid': 'Partiellement payée',
                'paid': 'Payée',
                'cancelled': 'Annulée'
            }
            
            data.append({
                'number': inv.number or f"INV-{inv.id}",
                'issue_date': inv.issue_date.strftime('%d/%m/%Y') if inv.issue_date else '',
                'due_date': inv.due_date.strftime('%d/%m/%Y') if inv.due_date else '',
                'client': client_name,
                'subtotal': subtotal,
                'tax_rate': tax_rate,
                'tax_amount': tax_amount,
                'total': total,
                'status': status_labels.get(inv.status, inv.status),
                'notes': inv.notes or ''
            })
        
        # Générer l'export selon le format
        if export_format == 'excel' and EXCEL_AVAILABLE:
            return self._generate_excel_export(data, start_date, end_date)
        else:
            return self._generate_csv_export(data, start_date, end_date)
    
    def _generate_csv_export(self, data, start_date, end_date):
        """Génère un export CSV."""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # En-têtes
        headers = [
            'Numéro Facture',
            'Date Émission',
            'Date Échéance',
            'Client',
            'Montant HT (€)',
            'TVA (%)',
            'TVA (€)',
            'Montant TTC (€)',
            'Statut',
            'Notes'
        ]
        writer.writerow(headers)
        
        # Données
        for row in data:
            writer.writerow([
                row['number'],
                row['issue_date'],
                row['due_date'],
                row['client'],
                f"{row['subtotal']:.2f}".replace('.', ','),
                f"{row['tax_rate']:.2f}".replace('.', ','),
                f"{row['tax_amount']:.2f}".replace('.', ','),
                f"{row['total']:.2f}".replace('.', ','),
                row['status'],
                row['notes']
            ])
        
        # Ligne de total
        if data:
            total_ht = sum(r['subtotal'] for r in data)
            total_tva = sum(r['tax_amount'] for r in data)
            total_ttc = sum(r['total'] for r in data)
            
            writer.writerow([])
            writer.writerow([
                'TOTAL',
                '',
                '',
                '',
                f"{total_ht:.2f}".replace('.', ','),
                '',
                f"{total_tva:.2f}".replace('.', ','),
                f"{total_ttc:.2f}".replace('.', ','),
                '',
                ''
            ])
        
        # Préparer la réponse
        output.seek(0)
        response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8-sig')
        filename = f"export_comptable_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def _generate_excel_export(self, data, start_date, end_date):
        """Génère un export Excel avec mise en forme."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export Comptable"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = [
            'Numéro Facture',
            'Date Émission',
            'Date Échéance',
            'Client',
            'Montant HT (€)',
            'TVA (%)',
            'TVA (€)',
            'Montant TTC (€)',
            'Statut',
            'Notes'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Données
        for row_num, row_data in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=row_data['number']).border = border
            ws.cell(row=row_num, column=2, value=row_data['issue_date']).border = border
            ws.cell(row=row_num, column=3, value=row_data['due_date']).border = border
            ws.cell(row=row_num, column=4, value=row_data['client']).border = border
            ws.cell(row=row_num, column=5, value=row_data['subtotal']).border = border
            ws.cell(row=row_num, column=6, value=row_data['tax_rate']).border = border
            ws.cell(row=row_num, column=7, value=row_data['tax_amount']).border = border
            ws.cell(row=row_num, column=8, value=row_data['total']).border = border
            ws.cell(row=row_num, column=9, value=row_data['status']).border = border
            ws.cell(row=row_num, column=10, value=row_data['notes']).border = border
            
            # Format numérique pour les montants
            for col in [5, 6, 7, 8]:
                ws.cell(row=row_num, column=col).number_format = '#,##0.00'
        
        # Ligne de total
        if data:
            total_row = len(data) + 3
            ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
            ws.cell(row=total_row, column=5, value=sum(r['subtotal'] for r in data)).font = Font(bold=True)
            ws.cell(row=total_row, column=5).number_format = '#,##0.00'
            ws.cell(row=total_row, column=7, value=sum(r['tax_amount'] for r in data)).font = Font(bold=True)
            ws.cell(row=total_row, column=7).number_format = '#,##0.00'
            ws.cell(row=total_row, column=8, value=sum(r['total'] for r in data)).font = Font(bold=True)
            ws.cell(row=total_row, column=8).number_format = '#,##0.00'
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 30
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Préparer la réponse
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"export_comptable_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @extend_schema(
        summary="Rapport des ventes",
        description="Rapport détaillé des ventes avec statistiques et évolution mensuelle.",
        parameters=[
            OpenApiParameter(name='start_date', description='Date début (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='end_date', description='Date fin (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='salesperson_id', description='ID du commercial (optionnel)', required=False, type=int),
        ]
    )
    @action(detail=False, methods=['get'], url_path='sales-report')
    def sales_report(self, request):
        """
        Rapport des ventes avec :
        - CA total
        - Nombre de devis signés
        - Ticket moyen
        - CA par commercial
        - Évolution mensuelle du CA
        """
        start_date, end_date = self._get_date_range(request)
        salesperson_id = request.query_params.get('salesperson_id')
        
        # Filtrer les devis signés (status='signed')
        # Quote -> Offer -> ProspectRequest (relation 'request')
        quotes = Quote.objects.filter(
            created_at__range=[start_date, end_date],
            status='accepted'
        ).select_related('offer__request__assigned_to')
        
        # Filtrer par commercial si demandé
        if salesperson_id:
            quotes = quotes.filter(offer__request__assigned_to_id=salesperson_id)
        
        # Statistiques globales
        total_revenue = quotes.aggregate(
            total=Coalesce(Sum('total'), Value(0), output_field=DecimalField())
        )['total']
        
        total_quotes = quotes.count()
        average_ticket = total_revenue / total_quotes if total_quotes > 0 else 0
        
        # CA par commercial
        revenue_by_salesperson = quotes.values(
            salesperson_id=F('offer__request__assigned_to__id'),
            salesperson_name=F('offer__request__assigned_to__last_name'),
            salesperson_firstname=F('offer__request__assigned_to__first_name')
        ).annotate(
            revenue=Coalesce(Sum('total'), Value(0), output_field=DecimalField()),
            quote_count=Count('id')
        ).order_by('-revenue')
        
        # Évolution mensuelle
        monthly_data = quotes.annotate(
            month=F('created_at__month'),
            year=F('created_at__year')
        ).values('year', 'month').annotate(
            revenue=Coalesce(Sum('total'), Value(0), output_field=DecimalField()),
            quote_count=Count('id')
        ).order_by('year', 'month')
        
        # Générer l'Excel
        return self._generate_sales_excel(
            total_revenue,
            total_quotes,
            average_ticket,
            revenue_by_salesperson,
            monthly_data,
            start_date,
            end_date
        )
    
    def _generate_sales_excel(self, total_revenue, total_quotes, average_ticket, 
                              revenue_by_salesperson, monthly_data, start_date, end_date):
        """Génère un rapport Excel des ventes."""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'openpyxl non disponible pour générer des fichiers Excel'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport des Ventes"
        
        # Styles
        title_font = Font(size=16, bold=True, color='1F4E78')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        kpi_font = Font(size=14, bold=True)
        
        # Titre
        ws['A1'] = 'RAPPORT DES VENTES'
        ws['A1'].font = title_font
        ws['A2'] = f"Période : {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        
        # KPIs
        row = 4
        ws[f'A{row}'] = 'Chiffre d\'affaires total'
        ws[f'B{row}'] = float(total_revenue)
        ws[f'B{row}'].number_format = '#,##0.00 €'
        ws[f'B{row}'].font = kpi_font
        
        row += 1
        ws[f'A{row}'] = 'Nombre de devis signés'
        ws[f'B{row}'] = total_quotes
        ws[f'B{row}'].font = kpi_font
        
        row += 1
        ws[f'A{row}'] = 'Ticket moyen'
        ws[f'B{row}'] = float(average_ticket)
        ws[f'B{row}'].number_format = '#,##0.00 €'
        ws[f'B{row}'].font = kpi_font
        
        # CA par commercial
        row += 3
        ws[f'A{row}'] = 'CHIFFRE D\'AFFAIRES PAR COMMERCIAL'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        headers = ['Commercial', 'CA (€)', 'Nb Devis', 'Ticket Moyen (€)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for salesperson in revenue_by_salesperson:
            row += 1
            name = f"{salesperson['salesperson_name']} {salesperson['salesperson_firstname']}" if salesperson['salesperson_name'] else "Sans commercial"
            revenue = float(salesperson['revenue'])
            count = salesperson['quote_count']
            avg = revenue / count if count > 0 else 0
            
            ws[f'A{row}'] = name
            ws[f'B{row}'] = revenue
            ws[f'B{row}'].number_format = '#,##0.00 €'
            ws[f'C{row}'] = count
            ws[f'D{row}'] = avg
            ws[f'D{row}'].number_format = '#,##0.00 €'
        
        # Évolution mensuelle
        row += 3
        ws[f'A{row}'] = 'ÉVOLUTION MENSUELLE'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        headers = ['Mois', 'CA (€)', 'Nb Devis', 'Ticket Moyen (€)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        months_fr = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin', 
                     'Juil', 'Août', 'Sep', 'Oct', 'Nov', 'Déc']
        
        for data in monthly_data:
            row += 1
            month_label = f"{months_fr[data['month']-1]} {data['year']}"
            revenue = float(data['revenue'])
            count = data['quote_count']
            avg = revenue / count if count > 0 else 0
            
            ws[f'A{row}'] = month_label
            ws[f'B{row}'] = revenue
            ws[f'B{row}'].number_format = '#,##0.00 €'
            ws[f'C{row}'] = count
            ws[f'D{row}'] = avg
            ws[f'D{row}'].number_format = '#,##0.00 €'
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        
        # Sauvegarder
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"rapport_ventes_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @extend_schema(
        summary="Rapport des commissions",
        description="Rapport détaillé des commissions à payer et payées.",
        parameters=[
            OpenApiParameter(name='start_date', description='Date début (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='end_date', description='Date fin (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='paid_status', description='Statut (all, paid, unpaid)', required=False, type=str),
        ]
    )
    @action(detail=False, methods=['get'], url_path='commissions-report')
    def commissions_report(self, request):
        """
        Rapport des commissions avec :
        - Liste des commissions dues/payées par personne
        - Totaux par type (apporteur d'affaires / commercial)
        - Détails par installation
        """
        start_date, end_date = self._get_date_range(request)
        paid_status = request.query_params.get('paid_status', 'all')
        
        # Récupérer les installations avec commissions
        installations = Form.objects.filter(
            created_at__range=[start_date, end_date]
        ).filter(
            Q(commission_amount__gt=0) | Q(sales_commission_amount__gt=0)
        ).select_related('client', 'offer__request__source', 'offer__request__assigned_to')
        
        # Filtrer par statut de paiement
        if paid_status == 'paid':
            installations = installations.filter(
                Q(commission_paid=True) | Q(sales_commission_paid=True)
            )
        elif paid_status == 'unpaid':
            installations = installations.filter(
                Q(commission_paid=False, commission_amount__gt=0) |
                Q(sales_commission_paid=False, sales_commission_amount__gt=0)
            )
        
        # Préparer les données
        commissions_data = []
        for installation in installations:
            client_name = installation.client.get_full_name() if installation.client else "Client inconnu"
            
            # Commission apporteur d'affaires
            if installation.commission_amount and installation.commission_amount > 0:
                source = installation.offer.request.source if hasattr(installation, 'offer') and installation.offer and hasattr(installation.offer, 'request') and installation.offer.request else None
                source_name = source.get_full_name() if source else "Source inconnue"
                
                commissions_data.append({
                    'type': 'Apporteur d\'affaires',
                    'beneficiary': source_name,
                    'client': client_name,
                    'amount': float(installation.commission_amount),
                    'paid': installation.commission_paid,
                    'installation_date': installation.created_at
                })
            
            # Commission commercial
            if installation.sales_commission_amount and installation.sales_commission_amount > 0:
                assigned_to = installation.offer.request.assigned_to if hasattr(installation, 'offer') and installation.offer and hasattr(installation.offer, 'request') and installation.offer.request else None
                commercial_name = assigned_to.get_full_name() if assigned_to else "Commercial inconnu"
                
                commissions_data.append({
                    'type': 'Commercial',
                    'beneficiary': commercial_name,
                    'client': client_name,
                    'amount': float(installation.sales_commission_amount),
                    'paid': installation.sales_commission_paid,
                    'installation_date': installation.created_at
                })
        
        # Générer l'Excel
        return self._generate_commissions_excel(commissions_data, start_date, end_date)
    
    def _generate_commissions_excel(self, commissions_data, start_date, end_date):
        """Génère un rapport Excel des commissions."""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'openpyxl non disponible pour générer des fichiers Excel'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport des Commissions"
        
        # Styles
        title_font = Font(size=16, bold=True, color='1F4E78')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        paid_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        unpaid_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Titre
        ws['A1'] = 'RAPPORT DES COMMISSIONS'
        ws['A1'].font = title_font
        ws['A2'] = f"Période : {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        
        # Statistiques globales
        total_commissions = sum(c['amount'] for c in commissions_data)
        total_paid = sum(c['amount'] for c in commissions_data if c['paid'])
        total_unpaid = total_commissions - total_paid
        
        row = 4
        ws[f'A{row}'] = 'Total commissions'
        ws[f'B{row}'] = total_commissions
        ws[f'B{row}'].number_format = '#,##0.00 €'
        ws[f'B{row}'].font = Font(size=12, bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Payées'
        ws[f'B{row}'] = total_paid
        ws[f'B{row}'].number_format = '#,##0.00 €'
        ws[f'B{row}'].fill = paid_fill
        
        row += 1
        ws[f'A{row}'] = 'À payer'
        ws[f'B{row}'] = total_unpaid
        ws[f'B{row}'].number_format = '#,##0.00 €'
        ws[f'B{row}'].fill = unpaid_fill
        
        # Liste détaillée
        row += 3
        ws[f'A{row}'] = 'DÉTAIL DES COMMISSIONS'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        headers = ['Type', 'Bénéficiaire', 'Client', 'Montant (€)', 'Statut', 'Date Installation']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Trier par bénéficiaire puis par type
        commissions_data_sorted = sorted(commissions_data, key=lambda x: (x['beneficiary'], x['type']))
        
        for commission in commissions_data_sorted:
            row += 1
            ws[f'A{row}'] = commission['type']
            ws[f'B{row}'] = commission['beneficiary']
            ws[f'C{row}'] = commission['client']
            ws[f'D{row}'] = commission['amount']
            ws[f'D{row}'].number_format = '#,##0.00 €'
            ws[f'E{row}'] = 'Payée' if commission['paid'] else 'À payer'
            ws[f'F{row}'] = commission['installation_date'].strftime('%d/%m/%Y')
            
            # Colorier selon le statut
            if commission['paid']:
                ws[f'E{row}'].fill = paid_fill
            else:
                ws[f'E{row}'].fill = unpaid_fill
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        
        # Sauvegarder
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"rapport_commissions_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @extend_schema(
        summary="Rapport des prospects",
        description="Rapport d'analyse du pipeline de prospects.",
        parameters=[
            OpenApiParameter(name='start_date', description='Date début (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='end_date', description='Date fin (YYYY-MM-DD)', required=False, type=str),
        ]
    )
    @action(detail=False, methods=['get'], url_path='prospects-report')
    def prospects_report(self, request):
        """
        Rapport des prospects avec :
        - Nombre de demandes par source
        - Taux de conversion par source
        - Temps moyen de traitement
        - Pipeline par statut
        """
        start_date, end_date = self._get_date_range(request)
        
        # Récupérer les demandes de la période
        prospects = ProspectRequest.objects.filter(
            created_at__range=[start_date, end_date]
        ).select_related('source', 'assigned_to')
        
        total_prospects = prospects.count()
        
        # Stats par source
        by_source = prospects.values('source_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Stats par statut
        by_status = prospects.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Taux de conversion (prospects -> devis signés)
        # Quote -> Offer -> ProspectRequest (relation 'request')
        signed_quotes = Quote.objects.filter(
            offer__request__in=prospects,
            status='signed'
        ).count()
        
        conversion_rate = (signed_quotes / total_prospects * 100) if total_prospects > 0 else 0
        
        # Génerer l'Excel
        return self._generate_prospects_excel(
            total_prospects,
            by_source,
            by_status,
            conversion_rate,
            start_date,
            end_date
        )
    
    def _generate_prospects_excel(self, total_prospects, by_source, by_status, 
                                   conversion_rate, start_date, end_date):
        """Génère un rapport Excel des prospects."""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'openpyxl non disponible pour générer des fichiers Excel'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport des Prospects"
        
        # Styles
        title_font = Font(size=16, bold=True, color='1F4E78')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        kpi_font = Font(size=14, bold=True)
        
        # Titre
        ws['A1'] = 'RAPPORT DES PROSPECTS'
        ws['A1'].font = title_font
        ws['A2'] = f"Période : {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        
        # KPIs
        row = 4
        ws[f'A{row}'] = 'Nombre total de prospects'
        ws[f'B{row}'] = total_prospects
        ws[f'B{row}'].font = kpi_font
        
        row += 1
        ws[f'A{row}'] = 'Taux de conversion'
        ws[f'B{row}'] = f"{conversion_rate:.1f}%"
        ws[f'B{row}'].font = kpi_font
        
        # Répartition par source
        row += 3
        ws[f'A{row}'] = 'RÉPARTITION PAR SOURCE'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        headers = ['Source', 'Nombre', 'Pourcentage']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        source_labels = {
            'website': 'Site Web',
            'referral': 'Recommandation',
            'social': 'Réseaux Sociaux',
            'direct': 'Contact Direct',
            'other': 'Autre'
        }
        
        for source in by_source:
            row += 1
            source_type = source['source_type'] or 'other'
            count = source['count']
            percentage = (count / total_prospects * 100) if total_prospects > 0 else 0
            
            ws[f'A{row}'] = source_labels.get(source_type, source_type)
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
        
        # Répartition par statut
        row += 3
        ws[f'A{row}'] = 'PIPELINE PAR STATUT'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        headers = ['Statut', 'Nombre', 'Pourcentage']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        status_labels = {
            'new': 'Nouveau',
            'contacted': 'Contacté',
            'qualified': 'Qualifié',
            'quote_sent': 'Devis envoyé',
            'won': 'Gagné',
            'lost': 'Perdu'
        }
        
        for status_data in by_status:
            row += 1
            status_type = status_data['status']
            count = status_data['count']
            percentage = (count / total_prospects * 100) if total_prospects > 0 else 0
            
            ws[f'A{row}'] = status_labels.get(status_type, status_type)
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Sauvegarder
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"rapport_prospects_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
